import openpyxl

class GetData():
    def __init__(self,fileName,sheetName):
        self.fileName = fileName
        self.sheetName = sheetName
        self.workbook = openpyxl.load_workbook(self.fileName)
    def get_rows(self):
        sheets = self.workbook[self.sheetName]
        return sheets.max_row
    def get_cols(self):
        sheets = self.workbook[self.sheetName]
        return sheets.max_column
    def read_data(self,rowNum,colNum):
        sheets = self.workbook[self.sheetName]
        return sheets.cell(row=rowNum,column=colNum).value
    def read_data_string(self,rowNum,colNum):
        sheets = self.workbook[self.sheetName]
        if str(sheets.cell(row = rowNum,column=colNum).value) == "None":
            return ""
        else:
            return str(sheets.cell(row=rowNum,column=colNum).value)
    def read_data_int(self,rowNum,colNum):
        sheets = self.workbook[self.sheetName]
        return int(sheets.cell(row=rowNum,column=colNum).value)
    def write_data(self,rowNum,colNum,Data):
        sheets = self.workbook[self.sheetName]
        sheets.cell(row=rowNum,column=colNum).value = Data
        self.workbook.save(self.fileName)


